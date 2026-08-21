from app.database import db
from app.models.employee import Employee
from datetime import datetime

from openpyxl import load_workbook


class EmployeeService:

    # ============================================================
    # CREATE EMPLOYEE
    # ============================================================

    @staticmethod
    def create_employee(data):

        required_fields = [
            "employee_id",
            "name",
            "email",
            "dob"
        ]

        for field in required_fields:

            if not data.get(field):

                return {
                    "success": False,
                    "message": f"{field} is required"
                }


        employee_id = data.get(
            "employee_id",
            ""
        ).strip()

        email = data.get(
            "email",
            ""
        ).strip()


        # Check Employee ID

        existing_employee = Employee.query.filter_by(
            employee_id=employee_id
        ).first()

        if existing_employee:

            return {
                "success": False,
                "message": "Employee ID already exists"
            }


        # Check Email

        existing_email = Employee.query.filter_by(
            email=email
        ).first()

        if existing_email:

            return {
                "success": False,
                "message": "Email already exists"
            }


        employee = Employee(

            employee_id=employee_id,

            name=data.get(
                "name"
            ).strip(),

            email=email,

            dob=datetime.strptime(
                data.get("dob"),
                "%Y-%m-%d"
            ).date(),

            department=data.get(
                "department",
                ""
            ).strip(),

            designation=data.get(
                "designation",
                ""
            ).strip()

        )


        db.session.add(employee)

        db.session.commit()


        return {
            "success": True,
            "message": "Employee created successfully"
        }


    # ============================================================
    # GET ALL EMPLOYEES
    # ============================================================

    @staticmethod
    def get_all_employees():

        employees = Employee.query.order_by(
            Employee.id.asc()
        ).all()

        employee_list = []


        for employee in employees:

            employee_list.append({

                "id": employee.id,

                "employee_id":
                    employee.employee_id,

                "name":
                    employee.name,

                "email":
                    employee.email,

                "dob":
                    employee.dob.strftime(
                        "%Y-%m-%d"
                    ),

                "department":
                    employee.department,

                "designation":
                    employee.designation,

                "status":
                    employee.status

            })


        return {
            "success": True,
            "data": employee_list
        }


    # ============================================================
    # UPDATE EMPLOYEE
    # ============================================================

    @staticmethod
    def update_employee(
        employee_id,
        data
    ):

        employee = Employee.query.get(
            employee_id
        )


        if not employee:

            return {
                "success": False,
                "message": "Employee not found"
            }


        employee.name = data.get(
            "name"
        )

        employee.email = data.get(
            "email"
        )

        employee.dob = datetime.strptime(
            data.get("dob"),
            "%Y-%m-%d"
        ).date()

        employee.department = data.get(
            "department"
        )

        employee.designation = data.get(
            "designation"
        )

        employee.status = data.get(
            "status",
            True
        )


        db.session.commit()


        return {
            "success": True,
            "message": "Employee updated successfully"
        }


    # ============================================================
    # DELETE EMPLOYEE
    # ============================================================

    @staticmethod
    def delete_employee(
        employee_id
    ):

        employee = Employee.query.get(
            employee_id
        )


        if not employee:

            return {
                "success": False,
                "message": "Employee not found"
            }


        db.session.delete(employee)

        db.session.commit()


        return {
            "success": True,
            "message": "Employee deleted successfully"
        }


    # ============================================================
    # IMPORT EMPLOYEES FROM EXCEL
    # ============================================================

    @staticmethod
    def import_from_excel(file):

        try:

            # ----------------------------------------------------
            # Load workbook
            # ----------------------------------------------------

            workbook = load_workbook(
                file,
                data_only=True
            )


            worksheet = workbook.active


            # ----------------------------------------------------
            # Check Excel headers
            # ----------------------------------------------------

            headers = []

            for cell in worksheet[1]:

                if cell.value is not None:

                    headers.append(
                        str(cell.value)
                        .strip()
                        .lower()
                    )


            required_headers = [
                "id",
                "name",
                "email",
                "dob"
            ]


            missing_headers = [

                header

                for header in required_headers

                if header not in headers

            ]


            if missing_headers:

                return {
                    "success": False,
                    "message":
                        "Missing required Excel columns.",
                    "missing_columns":
                        missing_headers
                }


            # ----------------------------------------------------
            # Map column names to indexes
            # ----------------------------------------------------

            header_map = {

                header: index

                for index, header
                in enumerate(headers)

            }


            # ----------------------------------------------------
            # Read Excel rows
            # ----------------------------------------------------

            rows = list(
                worksheet.iter_rows(
                    min_row=2,
                    values_only=True
                )
            )


            if not rows:

                return {
                    "success": False,
                    "message":
                        "Excel file contains no employee data."
                }


            # ----------------------------------------------------
            # Prepare import
            # ----------------------------------------------------

            employees_to_insert = []

            errors = []

            employee_ids_in_file = set()

            emails_in_file = set()


            # ----------------------------------------------------
            # Validate every row first
            # ----------------------------------------------------

            for row_number, row in enumerate(
                rows,
                start=2
            ):

                try:

                    employee_id = row[
                        header_map["id"]
                    ]

                    name = row[
                        header_map["name"]
                    ]

                    email = row[
                        header_map["email"]
                    ]

                    dob = row[
                        header_map["dob"]
                    ]


                    # Optional columns

                    department = ""

                    if "department" in header_map:

                        department = row[
                            header_map["department"]
                        ]


                    designation = ""

                    if "designation" in header_map:

                        designation = row[
                            header_map["designation"]
                        ]


                    status = True

                    if "status" in header_map:

                        excel_status = row[
                            header_map["status"]
                        ]

                        if excel_status is not None:

                            if isinstance(
                                excel_status,
                                bool
                            ):

                                status = excel_status

                            else:

                                status_text = str(
                                    excel_status
                                ).strip().lower()

                                status = (
                                    status_text
                                    not in [
                                        "false",
                                        "0",
                                        "inactive",
                                        "no"
                                    ]
                                )


                    # ------------------------------------------------
                    # Required value validation
                    # ------------------------------------------------

                    if employee_id is None:

                        raise ValueError(
                            "Employee ID is required"
                        )


                    if name is None:

                        raise ValueError(
                            "Name is required"
                        )


                    if email is None:

                        raise ValueError(
                            "Email is required"
                        )


                    if dob is None:

                        raise ValueError(
                            "DOB is required"
                        )


                    employee_id = str(
                        employee_id
                    ).strip()

                    name = str(
                        name
                    ).strip()

                    email = str(
                        email
                    ).strip()


                    if not employee_id:

                        raise ValueError(
                            "Employee ID is required"
                        )


                    if not name:

                        raise ValueError(
                            "Name is required"
                        )


                    if not email:

                        raise ValueError(
                            "Email is required"
                        )


                    # ------------------------------------------------
                    # Duplicate inside Excel
                    # ------------------------------------------------

                    employee_id_lower = (
                        employee_id.lower()
                    )

                    email_lower = (
                        email.lower()
                    )


                    if employee_id_lower in employee_ids_in_file:

                        raise ValueError(
                            "Duplicate Employee ID "
                            "inside Excel file"
                        )


                    if email_lower in emails_in_file:

                        raise ValueError(
                            "Duplicate email "
                            "inside Excel file"
                        )


                    employee_ids_in_file.add(
                        employee_id_lower
                    )

                    emails_in_file.add(
                        email_lower
                    )


                    # ------------------------------------------------
                    # Check database
                    # ------------------------------------------------

                    existing_employee = (
                        Employee.query.filter_by(
                            employee_id=employee_id
                        ).first()
                    )


                    if existing_employee:

                        raise ValueError(
                            f"Employee ID "
                            f"{employee_id} "
                            f"already exists"
                        )


                    existing_email = (
                        Employee.query.filter_by(
                            email=email
                        ).first()
                    )


                    if existing_email:

                        raise ValueError(
                            f"Email "
                            f"{email} "
                            f"already exists"
                        )


                    # ------------------------------------------------
                    # Convert DOB
                    # ------------------------------------------------

                    if isinstance(
                        dob,
                        datetime
                    ):

                        dob_value = dob.date()

                    elif hasattr(
                        dob,
                        "year"
                    ):

                        dob_value = dob

                    else:

                        dob_value = datetime.strptime(
                            str(dob).strip(),
                            "%Y-%m-%d"
                        ).date()


                    # ------------------------------------------------
                    # Optional values
                    # ------------------------------------------------

                    department = (
                        str(department).strip()
                        if department is not None
                        else ""
                    )


                    designation = (
                        str(designation).strip()
                        if designation is not None
                        else ""
                    )


                    # ------------------------------------------------
                    # Create employee object
                    # ------------------------------------------------

                    employee = Employee(

                        employee_id=employee_id,

                        name=name,

                        email=email,

                        dob=dob_value,

                        department=department,

                        designation=designation,

                        status=status

                    )


                    employees_to_insert.append(
                        employee
                    )


                except Exception as row_error:

                    errors.append({

                        "row": row_number,

                        "message": str(
                            row_error
                        )

                    })


            # ----------------------------------------------------
            # If validation errors exist
            # ----------------------------------------------------

            if errors:

                db.session.rollback()

                return {

                    "success": False,

                    "message":
                        "Excel import failed. "
                        "Please fix the errors and "
                        "upload the file again.",

                    "total_rows":
                        len(rows),

                    "inserted":
                        0,

                    "failed":
                        len(errors),

                    "errors":
                        errors

                }


            # ----------------------------------------------------
            # Insert all employees
            # ----------------------------------------------------

            db.session.add_all(
                employees_to_insert
            )

            db.session.commit()


            # ----------------------------------------------------
            # Success
            # ----------------------------------------------------

            return {

                "success": True,

                "message":
                    "Excel employees imported successfully.",

                "total_rows":
                    len(rows),

                "inserted":
                    len(employees_to_insert),

                "failed":
                    0,

                "errors":
                    []

            }


        except Exception as e:

            db.session.rollback()


            return {

                "success": False,

                "message":
                    "Excel import failed.",

                "error":
                    str(e)

            }